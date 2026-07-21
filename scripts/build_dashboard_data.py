"""
Reads the current month's channel sales files + the two Targets files from
Google Drive (via a service account) and produces data.json: a single file
the dashboard site can fetch and render, with target-vs-actual numbers per
channel and per SKU.

This script only ever READS from Drive -- it never writes back. The channel
"actual" numbers come from the same "Sales - <Month> <Year>" sheets that
Monthly_Target_Update_Code.py keeps up to date, so there's one source of
truth for what "actual" means.

Setup:
    pip install google-api-python-client google-auth openpyxl

Usage:
    python build_dashboard_data.py [path/to/service-account-key.json]
    (defaults to KEY_FILE below if no argument given)
"""

import calendar
import datetime as dt
import io
import json
import sys
from pathlib import Path

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# =============================================================================
# CONFIGURATION
# =============================================================================

KEY_FILE = "madeleine-targets-dashboard-00138fd3c845.json"
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# The "Monthly Targets" folder in Drive (found via check_drive_access.py).
# Anchoring on the ID rather than searching by name avoids ambiguity if
# multiple items ever share a name.
MONTHLY_TARGETS_FOLDER_ID = "1vcdwcPC9BmWdSMB51WhJsj_DrAGT5Bs9"

# Each run writes data/<year>-<month>.json (e.g. data/2026-07.json) rather
# than overwriting a single file, so past months stay available for the
# site's month/year filter. data/manifest.json is updated alongside it to
# list every month that has a file, newest first.
OUTPUT_DIR = Path("data")

# SKU -> public image URL, exported from Plytix. Kept as a static file
# bundled in the repo (see SKU Images.xlsx) rather than pulled from Drive.
SKU_IMAGES_PATH = Path("sku_images.json")

# Set to a specific date, e.g. dt.date(2026, 8, 5), to build the dashboard for
# a past month. Leave as None to use the real current month.
SIMULATED_TODAY = None

# --- Channel registry ---------------------------------------------------
# "file_stub": prefix used to find this channel's ACTUAL sales file in the
#   month folder (same matching convention as Monthly_Target_Update_Code.py).
# "target_sheet": sheet name to read for this channel's TARGET data, and
#   which Targets workbook ("MH EU" or "MH US") it lives in. None if there's
#   no target sheet for this channel yet.
CHANNEL_REGISTRY = [
    {"name": "Wayfair US", "region": "MH US", "file_stub": "Wayfair US",
     "target_workbook": "MH US", "target_sheet": "Wayfair US"},
    {"name": "Amazon US", "region": "MH US", "file_stub": "Amazon US",
     "target_workbook": "MH US", "target_sheet": "Amazon US"},
    {"name": "Wayfair EU", "region": "MH EU", "file_stub": "Wayfair EU",
     "target_workbook": "MH EU", "target_sheet": "Wayfair UK"},  # same channel, different label
    {"name": "Shopify UK", "region": "MH EU", "file_stub": "Shopify UK",
     "target_workbook": "MH EU", "target_sheet": "Shopify UK"},  # matched with .strip()
    {"name": "Debenhams", "region": "MH EU", "file_stub": "Debenhams",
     "target_workbook": "MH EU", "target_sheet": "Debenhams"},  # no actual file synced yet
    {"name": "Other Sales Channels", "region": "MH EU", "file_stub": "Amazon UK",
     "target_workbook": "MH EU", "target_sheet": "Other Sales Channels"},
    # -- channels with no target sheet yet; actual-only until Q4 targets land
    {"name": "Shopify US", "region": "MH US", "file_stub": "Shopify US", "target_workbook": None, "target_sheet": None},
    {"name": "Home Depot", "region": "MH US", "file_stub": "Home Depot", "target_workbook": None, "target_sheet": None},
    {"name": "Amazon EU", "region": "MH EU", "file_stub": "Amazon EU", "target_workbook": None, "target_sheet": None},
    {"name": "Overstock", "region": "MH US", "file_stub": "Overstock", "target_workbook": None, "target_sheet": None},
    {"name": "E-Bay UK", "region": "MH EU", "file_stub": "E-Bay UK", "target_workbook": None, "target_sheet": None},
    {"name": "Walmart", "region": "MH US", "file_stub": "Walmart", "target_workbook": None, "target_sheet": None},
    {"name": "Range", "region": "MH EU", "file_stub": "Range", "target_workbook": None, "target_sheet": None},
    {"name": "B&Q", "region": "MH EU", "file_stub": "B&Q", "target_workbook": None, "target_sheet": None},
    {"name": "Faire UK", "region": "MH EU", "file_stub": "Faire UK", "target_workbook": None, "target_sheet": None},
    {"name": "Lowes", "region": "MH US", "file_stub": "Lowes", "target_workbook": None, "target_sheet": None},
    {"name": "Faire", "region": "MH US", "file_stub": "Faire", "target_workbook": None, "target_sheet": None},
    {"name": "Houzz", "region": "MH US", "file_stub": "Houzz", "target_workbook": None, "target_sheet": None},
    {"name": "Tesco", "region": "MH EU", "file_stub": "Tesco", "target_workbook": None, "target_sheet": None},
]

TARGET_WORKBOOK_STUBS = {
    "MH EU": "MH EU - Revenue Targets",
    "MH US": "MH US - Wayfair & Amazon Monthly Targets",
}

# =============================================================================
# DRIVE HELPERS
# =============================================================================

def get_drive_service(key_file):
    creds = service_account.Credentials.from_service_account_file(key_file, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)

def list_children(service, parent_id):
    files, page_token = [], None
    while True:
        resp = service.files().list(
            q=f"'{parent_id}' in parents and trashed = false",
            fields="nextPageToken, files(id, name, mimeType)",
            pageSize=1000,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return files

def find_child_by_name(service, parent_id, name):
    for f in list_children(service, parent_id):
        if f["name"] == name:
            return f
    return None

def download_bytes(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf

# =============================================================================
# PARSING ACTUALS (from the "Sales - <Month> <Year>" sheet in each channel file)
# =============================================================================

def parse_actuals_sheet(file_bytes, sheet_name):
    """Returns {sku: {"units": int, "revenue": float}} summed across all rows."""
    wb = openpyxl.load_workbook(file_bytes, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))
        col_idx = {name: i for i, name in enumerate(headers) if name is not None}
        if "SKU" not in col_idx or "Quantity" not in col_idx or "Subtotal Sales" not in col_idx:
            return {}

        totals = {}
        for values in rows:
            if not values or all(v is None for v in values):
                continue
            sku = values[col_idx["SKU"]]
            if sku in (None, ""):
                continue
            sku = str(sku).strip()
            qty = values[col_idx["Quantity"]] or 0
            rev = values[col_idx["Subtotal Sales"]] or 0
            entry = totals.setdefault(sku, {"units": 0, "revenue": 0.0})
            entry["units"] += qty
            entry["revenue"] += rev
        return totals
    finally:
        wb.close()

# =============================================================================
# PARSING TARGETS
# =============================================================================

def parse_target_sheet(file_bytes, sheet_name):
    """Returns {sku: {"units": int, "revenue": float}} from a Targets sheet
    with columns SKU / Units To Be Sold (or 'Units to be sold') / Revenue Target."""
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    try:
        match = next((s for s in wb.sheetnames if s.strip() == sheet_name.strip()), None)
        if match is None:
            return {}
        ws = wb[match]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else h for h in next(rows)]
        col_idx = {name: i for i, name in enumerate(headers) if name is not None}
        # matches "Units To Be Sold" (MH EU sheets) and "Units to be sold" (MH US sheets)
        units_col = next((c for c in col_idx if c.lower() == "units to be sold"), None)
        revenue_col = next((c for c in col_idx if c.lower() == "revenue target"), None)
        sku_col = next((c for c in col_idx if c.lower() == "sku"), None)
        if sku_col is None or units_col is None or revenue_col is None:
            return {}

        totals = {}
        for values in rows:
            if not values or all(v is None for v in values):
                continue
            sku = values[col_idx[sku_col]]
            if sku in (None, ""):
                continue
            sku = str(sku).strip()
            units = values[col_idx[units_col]] or 0
            revenue = values[col_idx[revenue_col]] or 0
            entry = totals.setdefault(sku, {"units": 0, "revenue": 0.0})
            entry["units"] += units
            entry["revenue"] += revenue
        return totals
    finally:
        wb.close()

# =============================================================================
# FILE MATCHING (mirrors Monthly_Target_Update_Code.py's convention)
# =============================================================================

def match_by_stub(files, stub):
    stub_lower = stub.lower()
    candidates = [f for f in files if Path(f["name"]).stem.lower().startswith(stub_lower + " -")]
    if not candidates:
        return None
    return max(candidates, key=lambda f: len(stub))

def sheet_name_for(month, year):
    return f"Sales - {dt.date(year, month, 1).strftime('%B')} {year}"

# =============================================================================
# PACING / PROJECTION
# =============================================================================

def days_elapsed_and_total(month, year, today):
    """How many days of the month have actually happened, and how many days
    are in it. A past month counts as fully elapsed; the current month counts
    up to today; a future month (shouldn't normally happen) counts as zero."""
    days_in_month = calendar.monthrange(year, month)[1]
    if (year, month) < (today.year, today.month):
        return days_in_month, days_in_month
    if (year, month) > (today.year, today.month):
        return 0, days_in_month
    return min(today.day, days_in_month), days_in_month

def project_revenue(actual_revenue, days_elapsed, days_in_month):
    if days_elapsed <= 0:
        return 0.0
    return actual_revenue / days_elapsed * days_in_month

# =============================================================================
# SKU IMAGES
# =============================================================================

def load_sku_images():
    if not SKU_IMAGES_PATH.exists():
        return {}
    with open(SKU_IMAGES_PATH) as f:
        return json.load(f)

# =============================================================================
# MAIN BUILD
# =============================================================================

def build_dashboard_data(service, month, year, today=None, sku_images=None):
    today = today or dt.date.today()
    sku_images = sku_images if sku_images is not None else {}
    days_elapsed, days_in_month = days_elapsed_and_total(month, year, today)

    year_folder = find_child_by_name(service, MONTHLY_TARGETS_FOLDER_ID, str(year))
    if year_folder is None:
        raise RuntimeError(f"No '{year}' folder found under Monthly Targets.")
    month_name = dt.date(year, month, 1).strftime("%B")
    month_folder = find_child_by_name(service, year_folder["id"], month_name)
    if month_folder is None:
        raise RuntimeError(f"No '{month_name}' folder found under Monthly Targets/{year}.")

    files_in_folder = list_children(service, month_folder["id"])
    sheet_name = sheet_name_for(month, year)

    # Download + parse the two Targets workbooks once (only sheets we need,
    # looked up lazily per channel below).
    target_workbook_bytes = {}
    for label, stub in TARGET_WORKBOOK_STUBS.items():
        f = match_by_stub(files_in_folder, stub) or next(
            (x for x in files_in_folder if Path(x["name"]).stem.lower().startswith(stub.lower())), None
        )
        if f:
            target_workbook_bytes[label] = download_bytes(service, f["id"])

    channels_out = []
    for profile in CHANNEL_REGISTRY:
        actual_file = match_by_stub(files_in_folder, profile["file_stub"])
        actual_by_sku = {}
        if actual_file:
            file_bytes = download_bytes(service, actual_file["id"])
            actual_by_sku = parse_actuals_sheet(file_bytes, sheet_name)

        target_by_sku = {}
        if profile["target_workbook"] and profile["target_workbook"] in target_workbook_bytes:
            wb_bytes = target_workbook_bytes[profile["target_workbook"]]
            wb_bytes.seek(0)
            target_by_sku = parse_target_sheet(wb_bytes, profile["target_sheet"])

        if not actual_file and not target_by_sku:
            continue  # nothing to report for this channel this month

        all_skus = sorted(set(actual_by_sku) | set(target_by_sku))
        sku_rows = []
        for sku in all_skus:
            a = actual_by_sku.get(sku, {"units": 0, "revenue": 0.0})
            t = target_by_sku.get(sku, {"units": 0, "revenue": 0.0})
            sku_rows.append({
                "sku": sku,
                "target_units": t["units"],
                "target_revenue": round(t["revenue"], 2),
                "actual_units": a["units"],
                "actual_revenue": round(a["revenue"], 2),
                "image_url": sku_images.get(sku),
            })

        target_revenue = round(sum(r["target_revenue"] for r in sku_rows), 2)
        actual_revenue = round(sum(r["actual_revenue"] for r in sku_rows), 2)
        projected_revenue = round(project_revenue(actual_revenue, days_elapsed, days_in_month), 2)

        channels_out.append({
            "name": profile["name"],
            "region": profile["region"],
            "target_units": sum(r["target_units"] for r in sku_rows),
            "target_revenue": target_revenue,
            "actual_units": sum(r["actual_units"] for r in sku_rows),
            "actual_revenue": actual_revenue,
            "projected_revenue": projected_revenue,
            "has_actual_file": actual_file is not None,
            "has_target_data": bool(target_by_sku),
            "skus": sku_rows,
        })

    regions = {}
    for c in channels_out:
        r = regions.setdefault(c["region"], {
            "region": c["region"], "target_revenue": 0.0, "actual_revenue": 0.0,
            "target_units": 0, "actual_units": 0, "projected_revenue": 0.0,
        })
        r["target_revenue"] += c["target_revenue"]
        r["actual_revenue"] += c["actual_revenue"]
        r["target_units"] += c["target_units"]
        r["actual_units"] += c["actual_units"]
        r["projected_revenue"] += c["projected_revenue"]
    for r in regions.values():
        r["target_revenue"] = round(r["target_revenue"], 2)
        r["actual_revenue"] = round(r["actual_revenue"], 2)
        r["projected_revenue"] = round(r["projected_revenue"], 2)

    return {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "month": month,
        "year": year,
        "month_label": f"{month_name} {year}",
        "days_elapsed": days_elapsed,
        "days_in_month": days_in_month,
        "regions": sorted(regions.values(), key=lambda r: r["region"]),
        "channels": channels_out,
    }

def update_manifest(month, year, month_label):
    """Merge this month's entry into data/manifest.json, preserving whatever
    other months are already listed (past months aren't regenerated every
    run, so their manifest entries must survive)."""
    manifest_path = OUTPUT_DIR / "manifest.json"
    months = []
    if manifest_path.exists():
        with open(manifest_path) as f:
            months = json.load(f).get("months", [])

    months = [m for m in months if not (m["year"] == year and m["month"] == month)]
    months.append({"year": year, "month": month, "label": month_label})
    months.sort(key=lambda m: (m["year"], m["month"]), reverse=True)

    with open(manifest_path, "w") as f:
        json.dump({"months": months}, f, indent=2)

def main():
    key_file = sys.argv[1] if len(sys.argv) > 1 else KEY_FILE
    today = SIMULATED_TODAY or dt.date.today()

    print(f"Building dashboard data for {today.strftime('%B %Y')} ...")
    service = get_drive_service(key_file)
    sku_images = load_sku_images()
    print(f"  -> loaded {len(sku_images)} SKU image link(s) from {SKU_IMAGES_PATH}"
          if sku_images else f"  -> no {SKU_IMAGES_PATH} found; SKU images will be blank")

    data = build_dashboard_data(service, today.month, today.year, today=today, sku_images=sku_images)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{data['year']:04d}-{data['month']:02d}.json"
    with open(out_path, "w") as f:
        json.dump(data, f, indent=2)
    update_manifest(data["month"], data["year"], data["month_label"])

    print(f"Wrote {out_path}: {len(data['channels'])} channel(s).")
    for r in data["regions"]:
        print(f"  [{r['region']} total] target ${r['target_revenue']:,.2f} / actual ${r['actual_revenue']:,.2f}")
    for c in data["channels"]:
        flag = "" if c["has_actual_file"] else " (no actual file yet)"
        flag += "" if c["has_target_data"] else " (no target data yet)"
        print(f"  [{c['name']}] target ${c['target_revenue']:,.2f} / actual ${c['actual_revenue']:,.2f}{flag}")

if __name__ == "__main__":
    main()
