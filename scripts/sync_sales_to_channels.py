"""
Cloud version of Monthly_Target_Update_Code.py.

Reads Sales Master and syncs new PO+SKU rows into each channel's own workbook
in Drive -- the exact same append-only, dedupe-by-PO+SKU logic as the
original local script, just backed by the Drive API instead of a local
Google Drive Desktop sync folder, so this can run entirely inside GitHub
Actions with no dependency on any local machine being on.

REQUIRES: the service account must have EDITOR access to the "Monthly
Targets" folder in Drive, since this script writes new rows into the
per-channel workbooks that live there. If it only has Viewer access there,
every write will fail with a 403 permission error (visible in the Actions
log and in the error-notification email).

Sales Master is NEVER written to -- it is only ever opened in read-only mode
(see load_master_rows_for_periods below) and no save/upload call is ever made
against it. The "Sales Master - New" folder only needs VIEWER access; do not
grant it Editor. This is enforced both by the code (no write path exists for
it) and by keeping its Drive permission at Viewer, so a bug here couldn't
touch Sales Master even in principle.

Usage:
    python sync_sales_to_channels.py [path/to/service-account-key.json]
    (defaults to KEY_FILE below if no argument given)
"""

import datetime as dt
import io
import os
import smtplib
import sys
from email.message import EmailMessage
from pathlib import Path

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# =============================================================================
# CONFIGURATION
# =============================================================================

KEY_FILE = "madeleine-targets-dashboard-00138fd3c845.json"
SCOPES = ["https://www.googleapis.com/auth/drive"]  # read + WRITE

MONTHLY_TARGETS_FOLDER_ID = "1vcdwcPC9BmWdSMB51WhJsj_DrAGT5Bs9"
SALES_MASTER_FOLDER_ID = "1jFpIoVKEwIiW3p-mUjjIf3xIcrEFKRCT"
SALES_MASTER_SUBFOLDER_NAME = "Yearly Masters"
SALES_MASTER_SHEET = "Sheet1"

XLSX_MIMETYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Set this to a specific date, e.g. dt.date(2026, 8, 5), to simulate "today"
# for testing or backfills. Leave as None for normal day-to-day use.
SIMULATED_TODAY = None

# How many days into a new month we keep syncing BOTH the new month's files
# and the previous month's files, to catch late-arriving prior-month rows.
GRACE_CUTOFF_DAY = 11

# --- Error-notification email (same secrets as the dashboard build step) ---
EMAIL_ENABLED = True
EMAIL_FROM = os.environ.get("MONTHLY_TARGET_EMAIL_FROM", "")
EMAIL_APP_PASSWORD = os.environ.get("MONTHLY_TARGET_EMAIL_APP_PASSWORD", "")
EMAIL_TO = "operations25@madeleinehome.com"
EMAIL_SUBJECT_PREFIX = "[Monthly Target Update]"

# =============================================================================
# CHANNEL PROFILES -- identical to Monthly_Target_Update_Code.py
# =============================================================================

STANDARD_COLUMNS = [
    "P.O. Number",
    "PO Date",
    "SKU",
    "Quantity",
    "Wholesale Price",
    "Subtotal Sales",
    "Warehouse Name",
    "Warehouse Region",
    "Ship To City",
    "State New",
    "Zone",
    "Destination Country",
    "PO+SKU",
]

def _simple_profile(name, file_stub=None):
    return {
        "name": name,
        "file_stub": file_stub or name,
        "filter": {"type": "channel_equals", "sales_channel": name},
        "target_columns": STANDARD_COLUMNS,
    }

CHANNEL_PROFILES = [
    {
        "name": "Wayfair US",
        "file_stub": "Wayfair US",
        "filter": {"type": "channel_equals", "sales_channel": "Wayfair US"},
        "target_columns": [
            "P.O. Number",
            "PO Date",
            "SKU",
            "Quantity",
            "Wholesale Price",
            "Subtotal Sales",
            "Warehouse Name",
            "Ship To City",
            "State New",
            "Zone",
            "Warehouse Region",
            "PO+SKU",
        ],
    },
    _simple_profile("Wayfair EU"),
    _simple_profile("Amazon US"),
    _simple_profile("Shopify UK"),
    _simple_profile("Debenhams"),  # now auto-synced from Sales Master (has its own file)
    _simple_profile("Shopify US"),
    _simple_profile("Home Depot"),
    _simple_profile("Amazon EU"),
    _simple_profile("Overstock"),
    _simple_profile("E-Bay UK"),
    _simple_profile("Walmart"),
    _simple_profile("Range"),
    _simple_profile("B&Q"),
    _simple_profile("Faire UK"),
    _simple_profile("Lowes"),
    _simple_profile("Faire"),
    _simple_profile("Houzz"),
    _simple_profile("Tesco"),
]

CATCHALL_PROFILES = [
    {
        "name": "Amazon UK (other MH EU channels)",
        "file_stub": "Amazon UK",
        "is_catchall": True,
        "region": "MH EU",
        "always_exclude_channels": ["Debenhams", "Shopify UK", "Wayfair EU"],
        "target_columns": [
            "P.O. Number",
            "PO Date",
            "SKU",
            "Quantity",
            "Wholesale Price",
            "Subtotal Sales",
            "Warehouse Name",
            "Warehouse Region",
            "Ship To City",
            "State New",
            "Zone",
            "Destination Country",
            "PO+SKU",
            "Channel",
            "Category",
            "SKU Available in Target",
        ],
        "column_source_map": {"Channel": "Sales Channel"},
        "formula_templates": {
            "Category": "=XLOOKUP(C{row},Priority!A:A,Priority!B:B,0)",
            "SKU Available in Target": (
                "=XLOOKUP(C{row},'Target - Other Sales Channel'!A:A,"
                "'Target - Other Sales Channel'!A:A,0)"
            ),
        },
    },
]

ALL_PROFILES = CHANNEL_PROFILES + CATCHALL_PROFILES

# =============================================================================
# DATE / PERIOD LOGIC -- unchanged pure functions
# =============================================================================

def month_name(month, year):
    return dt.date(year, month, 1).strftime("%B")

def sheet_name_for(month, year):
    return f"Sales - {month_name(month, year)} {year}"

def previous_month(month, year):
    first_of_month = dt.date(year, month, 1)
    last_day_prev = first_of_month - dt.timedelta(days=1)
    return last_day_prev.month, last_day_prev.year

def active_periods(today):
    """Return a list of (month, year, is_grace_catchup) to process for `today`."""
    month, year, day = today.month, today.year, today.day
    prev_month, prev_year = previous_month(month, year)

    if day == 1:
        return [(prev_month, prev_year, True)]
    if day <= GRACE_CUTOFF_DAY:
        return [(month, year, False), (prev_month, prev_year, True)]
    return [(month, year, False)]

# =============================================================================
# DRIVE HELPERS
# =============================================================================

def get_drive_service(key_file):
    creds = service_account.Credentials.from_service_account_file(key_file, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)

def list_children(service, parent_id):
    files, page_token = [], None
    while True:
        resp = service.files().list(
            q=f"'{parent_id}' in parents and trashed = false",
            fields="nextPageToken, files(id, name, mimeType)",
            pageSize=1000,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return files

def find_child_by_name(service, parent_id, name):
    for f in list_children(service, parent_id):
        if f["name"] == name:
            return f
    return None

def download_bytes(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf

def upload_bytes(service, file_id, buf, mimetype=XLSX_MIMETYPE):
    """Overwrites an existing Drive file's content in place, keeping the same
    file ID (and therefore the same sharing/link/version history)."""
    media = MediaIoBaseUpload(buf, mimetype=mimetype, resumable=False)
    service.files().update(fileId=file_id, media_body=media).execute()

# =============================================================================
# SALES MASTER READING (via Drive instead of a local G:\ path)
# =============================================================================

def parse_po_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y"):
        try:
            return dt.datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None

def find_sales_master_file(service, year):
    yearly_folder = find_child_by_name(service, SALES_MASTER_FOLDER_ID, SALES_MASTER_SUBFOLDER_NAME)
    if yearly_folder is None:
        raise RuntimeError(f"No '{SALES_MASTER_SUBFOLDER_NAME}' folder found under Sales Master.")
    name = f"Sales Master {year}.xlsx"
    f = find_child_by_name(service, yearly_folder["id"], name)
    if f is None:
        raise RuntimeError(f"No '{name}' found under Sales Master/{SALES_MASTER_SUBFOLDER_NAME}.")
    return f

def load_master_rows_for_periods(service, periods):
    """Reads whichever yearly Sales Master file(s) are needed to cover every
    requested period (usually just one, but spans two if a grace-period
    catch-up crosses a Dec/Jan boundary), keeping rows whose PO Date falls in
    any requested (month, year). Each record is tagged with '_month'/'_year'.

    READ-ONLY: opens with read_only=True and never saves/uploads anything
    back to this file. Sales Master is only ever a data source here."""
    period_set = {(m, y) for m, y, _ in periods}
    years_needed = sorted({y for _, y, _ in periods})

    all_rows = []
    master_headers = None
    for year in years_needed:
        f = find_sales_master_file(service, year)
        file_bytes = download_bytes(service, f["id"])
        wb = openpyxl.load_workbook(file_bytes, data_only=True, read_only=True)
        try:
            ws = wb[SALES_MASTER_SHEET]
            rows = ws.iter_rows(values_only=True)
            headers = list(next(rows))
            col_idx = {name: i for i, name in enumerate(headers) if name is not None}

            if "PO Date" not in col_idx or "Sales Channel" not in col_idx:
                raise ValueError(
                    f"Sales Master sheet '{SALES_MASTER_SHEET}' in {f['name']} must contain "
                    f"'PO Date' and 'Sales Channel' columns. Found columns: {headers}"
                )
            if master_headers is None:
                master_headers = set(col_idx.keys())

            for values in rows:
                if not values or all(v is None for v in values):
                    continue
                po_date = parse_po_date(values[col_idx["PO Date"]])
                if po_date is None or (po_date.month, po_date.year) not in period_set:
                    continue
                record = {name: values[idx] for name, idx in col_idx.items()}
                record["_month"] = po_date.month
                record["_year"] = po_date.year
                all_rows.append(record)
        finally:
            wb.close()

    return all_rows, (master_headers or set())

def channel_matches(record, filter_cfg):
    ftype = filter_cfg["type"]
    if ftype == "channel_equals":
        channel = record.get("Sales Channel")
        return channel is not None and str(channel).strip() == filter_cfg["sales_channel"]
    if ftype == "region_exclude_channels":
        region = record.get("Region")
        if region is None or str(region).strip() != filter_cfg["region"]:
            return False
        channel = record.get("Sales Channel")
        channel = str(channel).strip() if channel is not None else ""
        return channel not in filter_cfg["exclude_channels"]
    raise ValueError(f"Unknown filter type: {ftype}")

def project_channel_records(master_rows, master_headers, channel_cfg):
    formula_columns = list(channel_cfg.get("formula_templates", {}).keys())
    data_columns = [c for c in channel_cfg["target_columns"] if c not in formula_columns]
    source_map = channel_cfg.get("column_source_map", {})

    filter_cfg = channel_cfg["filter"]
    required = ["Sales Channel"]
    if filter_cfg["type"] == "region_exclude_channels":
        required.append("Region")
    required += [source_map.get(c, c) for c in data_columns]
    missing = [c for c in required if c not in master_headers]
    if missing:
        raise ValueError(
            f"Sales Master is missing column(s) required for channel "
            f"'{channel_cfg['name']}': {missing}"
        )

    seen_keys = set()
    matched = []
    for record in master_rows:
        if not channel_matches(record, filter_cfg):
            continue

        projected = {col: record.get(source_map.get(col, col)) for col in data_columns}

        key = projected.get("PO+SKU")
        key = str(key).strip() if key not in (None, "") else ""
        if key:
            if key in seen_keys:
                continue  # duplicate PO+SKU within Sales Master itself
            seen_keys.add(key)

        matched.append(projected)

    return matched

# =============================================================================
# FOLDER DISCOVERY / PROFILE MATCHING (Drive listing instead of local fs)
# =============================================================================

IGNORED_FILE_STUBS = [
    "MH EU - Revenue Targets",
    "MH US - Wayfair & Amazon Monthly Targets",
    "SKU Details and Priority",
]

def is_ignored_file(name):
    stem_lower = Path(name).stem.lower()
    return any(stem_lower.startswith(stub.lower()) for stub in IGNORED_FILE_STUBS)

def match_file_to_profile(name, profiles):
    stem_lower = Path(name).stem.lower()
    candidates = [p for p in profiles if stem_lower.startswith(p["file_stub"].lower() + " -")]
    if not candidates:
        return None
    return max(candidates, key=lambda p: len(p["file_stub"]))

def discover_and_match(service, month, year):
    """Returns (matched, unmatched_names).
    matched: {profile_name: (drive_file_dict, profile)}
    unmatched_names: [file_name, ...] -- present in the folder but no profile matches."""
    year_folder = find_child_by_name(service, MONTHLY_TARGETS_FOLDER_ID, str(year))
    if year_folder is None:
        return {}, []
    month_folder = find_child_by_name(service, year_folder["id"], month_name(month, year))
    if month_folder is None:
        return {}, []

    matched = {}
    unmatched_names = []
    for f in list_children(service, month_folder["id"]):
        if f["mimeType"] == "application/vnd.google-apps.folder":
            continue
        if f["name"].startswith("~$") or not f["name"].lower().endswith(".xlsx"):
            continue
        if is_ignored_file(f["name"]):
            continue
        profile = match_file_to_profile(f["name"], ALL_PROFILES)
        if profile is None:
            unmatched_names.append(f["name"])
        else:
            matched[profile["name"]] = (f, profile)
    return matched, unmatched_names

def resolve_catchall_filter(catchall_cfg, channel_names_with_own_file):
    exclude = set(catchall_cfg["always_exclude_channels"]) | set(channel_names_with_own_file)
    return {
        "type": "region_exclude_channels",
        "region": catchall_cfg["region"],
        "exclude_channels": list(exclude),
    }

# =============================================================================
# SYNCING A SINGLE CHANNEL FILE (append-only, never clears existing rows)
# =============================================================================

def get_last_data_row(ws, n_cols):
    last_row = 1  # header
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, n_cols + 1)):
            last_row = r
    return last_row

def sync_append_dedupe(ws, records, target_columns, formula_templates=None):
    """Never clears existing rows. For every record, checks whether its PO+SKU
    key is already present anywhere in the sheet; if not, appends it as a new
    row. Any column listed in formula_templates gets its formula written
    (with the row number substituted in), only for newly appended rows."""
    formula_templates = formula_templates or {}

    po_sku_idx = target_columns.index("PO+SKU")
    existing = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=po_sku_idx + 1).value
        if val not in (None, ""):
            existing.add(str(val).strip())

    next_row = get_last_data_row(ws, len(target_columns)) + 1

    added, skipped_dupe, skipped_no_key = 0, 0, 0
    for record in records:
        key = record.get("PO+SKU")
        key = str(key).strip() if key not in (None, "") else ""
        if not key:
            skipped_no_key += 1
            continue
        if key in existing:
            skipped_dupe += 1
            continue

        for i, col_name in enumerate(target_columns, start=1):
            if col_name in formula_templates:
                ws.cell(row=next_row, column=i).value = formula_templates[col_name].format(row=next_row)
            else:
                ws.cell(row=next_row, column=i).value = record.get(col_name)

        existing.add(key)
        next_row += 1
        added += 1

    summary = f"added {added} new row(s), skipped {skipped_dupe} duplicate(s)"
    if skipped_no_key:
        summary += f", skipped {skipped_no_key} row(s) with a blank PO+SKU"
    if formula_templates and added:
        summary += f", wrote formula(s) for {added} new row(s) in {', '.join(formula_templates)}"
    return summary

def sync_channel(service, channel_cfg, drive_file, sheet_name, records):
    target_columns = channel_cfg["target_columns"]
    formula_templates = channel_cfg.get("formula_templates", {})

    file_bytes = download_bytes(service, drive_file["id"])
    wb = openpyxl.load_workbook(file_bytes)  # formulas preserved (data_only not set)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {drive_file['name']}")
    ws = wb[sheet_name]

    for i, col_name in enumerate(target_columns, start=1):
        ws.cell(row=1, column=i).value = col_name

    summary = sync_append_dedupe(ws, records, target_columns, formula_templates)

    out_buf = io.BytesIO()
    wb.save(out_buf)
    wb.close()
    out_buf.seek(0)
    upload_bytes(service, drive_file["id"], out_buf)
    return summary

# =============================================================================
# ERROR EMAIL
# =============================================================================

def send_error_email(problems):
    if not EMAIL_ENABLED or not problems:
        return
    if not EMAIL_FROM or not EMAIL_APP_PASSWORD:
        print(
            "\n(!) Skipping error email: MONTHLY_TARGET_EMAIL_FROM / "
            "MONTHLY_TARGET_EMAIL_APP_PASSWORD are not set."
        )
        return

    body = "The Monthly Target Update run hit the following issue(s):\n\n" + "\n".join(
        f"- {p}" for p in problems
    )

    msg = EmailMessage()
    msg["Subject"] = f"{EMAIL_SUBJECT_PREFIX} {len(problems)} issue(s) on {dt.date.today().isoformat()}"
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.set_content(body)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL_FROM, EMAIL_APP_PASSWORD)
            server.send_message(msg)
        print(f"\n-> error email sent to {EMAIL_TO} ({len(problems)} issue(s)).")
    except Exception as exc:
        print(f"\n(!) FAILED to send error email: {exc}")

# =============================================================================
# MAIN
# =============================================================================

def process_period(service, month, year, all_master_rows, master_headers, is_grace_catchup, problems):
    label = f"{month_name(month, year)} {year}" + (" (grace catch-up)" if is_grace_catchup else "")
    print(f"\n--- {label} ---")

    matched, unmatched_names = discover_and_match(service, month, year)

    if not matched and not unmatched_names:
        print("  no files found in this month's Drive folder -- nothing to do.")
        return

    for name in unmatched_names:
        msg = f"[{label}] Found file '{name}' with no matching channel profile -> skipped."
        print(f"  {msg}")
        problems.append(msg)

    channel_names_with_own_file = {
        pname for pname, (_f, profile) in matched.items() if not profile.get("is_catchall")
    }

    master_rows = [r for r in all_master_rows if (r["_month"], r["_year"]) == (month, year)]
    sheet_name = sheet_name_for(month, year)

    for profile_name, (drive_file, profile) in matched.items():
        try:
            if profile.get("is_catchall"):
                filter_cfg = resolve_catchall_filter(
                    profile, channel_names_with_own_file - {profile_name}
                )
                channel_cfg = {**profile, "filter": filter_cfg}
            else:
                channel_cfg = profile

            records = project_channel_records(master_rows, master_headers, channel_cfg)
            summary = sync_channel(service, channel_cfg, drive_file, sheet_name, records)
            print(f"  [{profile_name}] {len(records)} matching row(s) -> {summary}.")
        except Exception as exc:
            msg = f"[{label}] {profile_name} FAILED: {exc}"
            print(f"  {msg}")
            problems.append(msg)

def main():
    key_file = sys.argv[1] if len(sys.argv) > 1 else KEY_FILE
    problems = []
    today = SIMULATED_TODAY or dt.date.today()
    periods = active_periods(today)

    print(f"Today: {today.isoformat()}  |  Active period(s): "
          + ", ".join(f"{month_name(m, y)} {y}{' (grace)' if g else ''}" for m, y, g in periods))

    service = get_drive_service(key_file)

    try:
        all_master_rows, master_headers = load_master_rows_for_periods(service, periods)
    except Exception as exc:
        msg = f"FATAL: could not read Sales Master: {exc}"
        print(msg)
        send_error_email([msg])
        return

    print(f"  -> {len(all_master_rows)} Sales Master row(s) loaded across active period(s).")

    for month, year, is_grace_catchup in periods:
        process_period(service, month, year, all_master_rows, master_headers, is_grace_catchup, problems)

    if problems:
        print(f"\n{len(problems)} issue(s) encountered this run:")
        for p in problems:
            print(f"  - {p}")
    send_error_email(problems)

if __name__ == "__main__":
    main()
