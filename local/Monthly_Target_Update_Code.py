import datetime as dt
import os
import smtplib
from email.message import EmailMessage
from pathlib import Path

import openpyxl

# =============================================================================
# CONFIGURATION
# =============================================================================

MONTHLY_TARGETS_ROOT = Path(r"G:\My Drive\Monthly Targets")

SALES_MASTER_PATH = Path(
    r"G:\.shortcut-targets-by-id\1jFpIoVKEwIiW3p-mUjjIf3xIcrEFKRCT\Sales Master - New\Yearly Masters\Sales Master 2026.xlsx"
)
SALES_MASTER_SHEET = "Sheet1"

# Set this to a specific date, e.g. dt.date(2026, 8, 5), to simulate "today" for
# testing or backfills. Leave as None for normal day-to-day use (uses the real
# current date).
SIMULATED_TODAY = None

# How many days into a new month we keep syncing BOTH the new month's files and
# the previous month's files, to catch late-arriving prior-month rows. Day 1 of
# a new month always processes the previous month ONLY (the Sales Master has no
# new-month data yet). Days 2..GRACE_CUTOFF_DAY process both months. From
# GRACE_CUTOFF_DAY+1 onward, only the current month is processed.
GRACE_CUTOFF_DAY = 11

# --- Error-notification email -----------------------------------------------
# Sent ONLY when something goes wrong (a channel fails, or a file in a month
# folder doesn't match any registered profile) -- never on a clean run.
#
# Credentials are read from environment variables rather than hardcoded here,
# because this file is headed for a public/shared GitHub repo -- never put a
# real password directly in this script.
#
# One-time setup (Gmail with 2-Step Verification enabled):
#   1. Go to https://myaccount.google.com/apppasswords
#   2. Generate an App Password for "Mail".
#   3. Set two environment variables on the machine that runs this script:
#        MONTHLY_TARGET_EMAIL_FROM = the Gmail address sending the alert
#        MONTHLY_TARGET_EMAIL_APP_PASSWORD = the 16-character app password
#      On Windows (Command Prompt, run once, then open a NEW terminal):
#        setx MONTHLY_TARGET_EMAIL_FROM "youraddress@gmail.com"
#        setx MONTHLY_TARGET_EMAIL_APP_PASSWORD "xxxxxxxxxxxxxxxx"
#   (When this later moves to GitHub Actions, the same two names get added as
#   encrypted repository Secrets instead of environment variables.)
EMAIL_ENABLED = True
EMAIL_FROM = os.environ.get("MONTHLY_TARGET_EMAIL_FROM", "")
EMAIL_APP_PASSWORD = os.environ.get("MONTHLY_TARGET_EMAIL_APP_PASSWORD", "")
EMAIL_TO = "operations25@madeleinehome.com"
EMAIL_SUBJECT_PREFIX = "[Monthly Target Update]"

# =============================================================================
# CHANNEL PROFILES
# One entry per Sales Channel value that can get its own dedicated file.
# "file_stub" is matched as a filename prefix against whatever is actually
# found in a given month's folder (e.g. stub "Wayfair US" matches
# "Wayfair US - October 2026.xlsx"). No path is stored -- the real file found
# on disk is always used, whatever its exact name turns out to be.
#
# NOTE: columns for the 13 channels below that don't have a dedicated file yet
# (everything except Wayfair US/EU, Amazon US, Shopify UK) are a best-guess
# default (STANDARD_COLUMNS, matching the majority pattern of the existing
# simple channels). Double check / adjust target_columns for each one against
# its real header row the first time a dedicated file for it shows up.
# =============================================================================

STANDARD_COLUMNS = [
    "P.O. Number",
    "PO Date",
    "SKU",
    "Quantity",
    "Wholesale Price",
    "Subtotal Sales",
    "Warehouse Name",
    "Warehouse Region",
    "Ship To City",
    "State New",
    "Zone",
    "Destination Country",
    "PO+SKU",
]

def _simple_profile(name, file_stub=None):
    return {
        "name": name,
        "file_stub": file_stub or name,
        "filter": {"type": "channel_equals", "sales_channel": name},
        "target_columns": STANDARD_COLUMNS,
    }

CHANNEL_PROFILES = [
    {
        "name": "Wayfair US",
        "file_stub": "Wayfair US",
        "filter": {"type": "channel_equals", "sales_channel": "Wayfair US"},
        "target_columns": [
            "P.O. Number",
            "PO Date",
            "SKU",
            "Quantity",
            "Wholesale Price",
            "Subtotal Sales",
            "Warehouse Name",
            "Ship To City",
            "State New",
            "Zone",
            "Warehouse Region",
            "PO+SKU",
        ],
    },
    _simple_profile("Wayfair EU"),
    _simple_profile("Amazon US"),
    _simple_profile("Shopify UK"),
    # -- channels below don't have a dedicated file yet as of Q3 2026; will be
    #    picked up automatically the moment a file for them appears in a month
    #    folder. See the docstring note above about verifying target_columns.
    _simple_profile("Shopify US"),
    _simple_profile("Home Depot"),
    _simple_profile("Amazon EU"),
    _simple_profile("Overstock"),
    _simple_profile("E-Bay UK"),
    _simple_profile("Walmart"),
    _simple_profile("Range"),
    _simple_profile("B&Q"),
    _simple_profile("Faire UK"),
    _simple_profile("Lowes"),
    _simple_profile("Faire"),
    _simple_profile("Houzz"),
    _simple_profile("Tesco"),
]

# The historical "leftover bucket" file for MH EU channels that don't (yet)
# have their own dedicated file. always_exclude_channels covers channels that
# are NEVER this script's responsibility (Debenhams is handled by a separate
# process entirely). At runtime this is further extended with whichever other
# channels turn out to have their own dedicated file present in that month's
# folder, so nothing gets double-counted once a channel graduates to its own
# file -- no manual list maintenance needed when that happens.
CATCHALL_PROFILES = [
    {
        "name": "Amazon UK (other MH EU channels)",
        "file_stub": "Amazon UK",
        "is_catchall": True,
        "region": "MH EU",
        "always_exclude_channels": ["Debenhams", "Shopify UK", "Wayfair EU"],
        "target_columns": [
            "P.O. Number",
            "PO Date",
            "SKU",
            "Quantity",
            "Wholesale Price",
            "Subtotal Sales",
            "Warehouse Name",
            "Warehouse Region",
            "Ship To City",
            "State New",
            "Zone",
            "Destination Country",
            "PO+SKU",
            "Channel",
            "Category",
            "SKU Available in Target",
        ],
        "column_source_map": {"Channel": "Sales Channel"},
        # Column O ("Category") and column P ("SKU Available in Target") are
        # formulas. Every time a NEW row is appended, the formula is written
        # with its cell reference re-pointed at that row's own "SKU" cell
        # (column C) -- e.g. row 53 gets C53, row 102 gets C102, etc.
        "formula_templates": {
            "Category": "=XLOOKUP(C{row},Priority!A:A,Priority!B:B,0)",
            "SKU Available in Target": (
                "=XLOOKUP(C{row},'Target - Other Sales Channel'!A:A,"
                "'Target - Other Sales Channel'!A:A,0)"
            ),
        },
    },
]

ALL_PROFILES = CHANNEL_PROFILES + CATCHALL_PROFILES

# =============================================================================
# DATE / PERIOD LOGIC
# =============================================================================

def month_name(month, year):
    return dt.date(year, month, 1).strftime("%B")

def sheet_name_for(month, year):
    return f"Sales - {month_name(month, year)} {year}"

def month_folder(month, year):
    return MONTHLY_TARGETS_ROOT / str(year) / month_name(month, year)

def previous_month(month, year):
    first_of_month = dt.date(year, month, 1)
    last_day_prev = first_of_month - dt.timedelta(days=1)
    return last_day_prev.month, last_day_prev.year

def active_periods(today):
    """Return a list of (month, year, is_grace_catchup) to process for `today`."""
    month, year, day = today.month, today.year, today.day
    prev_month, prev_year = previous_month(month, year)

    if day == 1:
        return [(prev_month, prev_year, True)]
    if day <= GRACE_CUTOFF_DAY:
        return [(month, year, False), (prev_month, prev_year, True)]
    return [(month, year, False)]

# =============================================================================
# SALES MASTER READING
# =============================================================================

def parse_po_date(value):
    """Return a date/datetime for month/year filtering, or None if unparseable."""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value
    if isinstance(value, dt.date):
        return dt.datetime(value.year, value.month, value.day)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y"):
        try:
            return dt.datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None

def load_master_rows_for_periods(periods):
    """Single read of the Sales Master, keeping rows whose PO Date falls in any
    of the requested (month, year) periods. Each returned record is tagged with
    '_month'/'_year' so callers can split it back out per period."""
    period_set = {(m, y) for m, y, _ in periods}

    wb = openpyxl.load_workbook(SALES_MASTER_PATH, data_only=True, read_only=True)
    try:
        ws = wb[SALES_MASTER_SHEET]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))
        col_idx = {name: i for i, name in enumerate(headers) if name is not None}

        if "PO Date" not in col_idx or "Sales Channel" not in col_idx:
            raise ValueError(
                f"Sales Master sheet '{SALES_MASTER_SHEET}' must contain 'PO Date' "
                f"and 'Sales Channel' columns. Found columns: {headers}"
            )

        all_rows = []
        for values in rows:
            if not values or all(v is None for v in values):
                continue
            po_date = parse_po_date(values[col_idx["PO Date"]])
            if po_date is None or (po_date.month, po_date.year) not in period_set:
                continue
            record = {name: values[idx] for name, idx in col_idx.items()}
            record["_month"] = po_date.month
            record["_year"] = po_date.year
            all_rows.append(record)

        return all_rows, set(col_idx.keys())
    finally:
        wb.close()

def channel_matches(record, filter_cfg):
    ftype = filter_cfg["type"]
    if ftype == "channel_equals":
        channel = record.get("Sales Channel")
        return channel is not None and str(channel).strip() == filter_cfg["sales_channel"]
    if ftype == "region_exclude_channels":
        region = record.get("Region")
        if region is None or str(region).strip() != filter_cfg["region"]:
            return False
        channel = record.get("Sales Channel")
        channel = str(channel).strip() if channel is not None else ""
        return channel not in filter_cfg["exclude_channels"]
    raise ValueError(f"Unknown filter type: {ftype}")

def project_channel_records(master_rows, master_headers, channel_cfg):
    formula_columns = list(channel_cfg.get("formula_templates", {}).keys())
    data_columns = [c for c in channel_cfg["target_columns"] if c not in formula_columns]
    source_map = channel_cfg.get("column_source_map", {})

    filter_cfg = channel_cfg["filter"]
    required = ["Sales Channel"]
    if filter_cfg["type"] == "region_exclude_channels":
        required.append("Region")
    required += [source_map.get(c, c) for c in data_columns]
    missing = [c for c in required if c not in master_headers]
    if missing:
        raise ValueError(
            f"Sales Master is missing column(s) required for channel "
            f"'{channel_cfg['name']}': {missing}"
        )

    seen_keys = set()
    matched = []
    for record in master_rows:
        if not channel_matches(record, filter_cfg):
            continue

        projected = {col: record.get(source_map.get(col, col)) for col in data_columns}

        key = projected.get("PO+SKU")
        key = str(key).strip() if key not in (None, "") else ""
        if key:
            if key in seen_keys:
                continue  # duplicate PO+SKU within Sales Master itself
            seen_keys.add(key)

        matched.append(projected)

    return matched

# =============================================================================
# FOLDER DISCOVERY / PROFILE MATCHING
# =============================================================================

# Files that legitimately live in the same month folder but are NOT per-channel
# sales-paste files (target trackers, lookup sheets, etc.) -- matched the same
# way as a channel stub (filename prefix, month/year varies). These are
# skipped quietly: never logged as a problem, never emailed.
IGNORED_FILE_STUBS = [
    "MH EU - Revenue Targets",
    "MH US - Wayfair & Amazon Monthly Targets",
    "SKU Details and Priority",
]

def is_ignored_file(file_path):
    stem_lower = file_path.stem.lower()
    return any(stem_lower.startswith(stub.lower()) for stub in IGNORED_FILE_STUBS)

def list_excel_files(folder):
    if not folder.exists():
        return []
    files = []
    for p in sorted(folder.iterdir()):
        if p.suffix.lower() != ".xlsx":
            continue
        if p.name.startswith("~$"):
            continue  # Excel's temporary lock file
        files.append(p)
    return files

def match_file_to_profile(file_path, profiles):
    """Return the profile whose file_stub is a '<stub> - ...' prefix of this
    file's name, or None if nothing matches. Longest stub wins on overlap."""
    stem_lower = file_path.stem.lower()
    candidates = [p for p in profiles if stem_lower.startswith(p["file_stub"].lower() + " -")]
    if not candidates:
        return None
    return max(candidates, key=lambda p: len(p["file_stub"]))

def discover_and_match(folder):
    """Returns (matched, unmatched_files).
    matched: {profile_name: (file_path, profile)}
    unmatched_files: [file_path, ...] -- present in the folder but no profile matches."""
    matched = {}
    unmatched_files = []
    for f in list_excel_files(folder):
        if is_ignored_file(f):
            continue
        profile = match_file_to_profile(f, ALL_PROFILES)
        if profile is None:
            unmatched_files.append(f)
        else:
            matched[profile["name"]] = (f, profile)
    return matched, unmatched_files

def resolve_catchall_filter(catchall_cfg, channel_names_with_own_file):
    exclude = set(catchall_cfg["always_exclude_channels"]) | set(channel_names_with_own_file)
    return {
        "type": "region_exclude_channels",
        "region": catchall_cfg["region"],
        "exclude_channels": list(exclude),
    }

# =============================================================================
# SYNCING A SINGLE CHANNEL FILE (append-only, never clears existing rows)
# =============================================================================

def get_last_data_row(ws, n_cols):
    """Find the last row (>=1) that actually has data in the first n_cols columns."""
    last_row = 1  # header
    for r in range(2, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, n_cols + 1)):
            last_row = r
    return last_row

def sync_append_dedupe(ws, records, target_columns, formula_templates=None):
    """Never clears existing rows. For every record, checks whether its PO+SKU
    key is already present anywhere in the sheet; if not, appends it as a new
    row. Any column listed in formula_templates gets its formula written (with
    the row number substituted in) instead of a plain value, only for rows
    that are newly appended."""
    formula_templates = formula_templates or {}

    po_sku_idx = target_columns.index("PO+SKU")  # 0-based
    existing = set()
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=po_sku_idx + 1).value
        if val not in (None, ""):
            existing.add(str(val).strip())

    next_row = get_last_data_row(ws, len(target_columns)) + 1

    added, skipped_dupe, skipped_no_key = 0, 0, 0
    for record in records:
        key = record.get("PO+SKU")
        key = str(key).strip() if key not in (None, "") else ""
        if not key:
            skipped_no_key += 1
            continue
        if key in existing:
            skipped_dupe += 1
            continue

        for i, col_name in enumerate(target_columns, start=1):
            if col_name in formula_templates:
                ws.cell(row=next_row, column=i).value = formula_templates[col_name].format(row=next_row)
            else:
                ws.cell(row=next_row, column=i).value = record.get(col_name)

        existing.add(key)
        next_row += 1
        added += 1

    summary = f"added {added} new row(s), skipped {skipped_dupe} duplicate(s)"
    if skipped_no_key:
        summary += f", skipped {skipped_no_key} row(s) with a blank PO+SKU"
    if formula_templates and added:
        summary += f", wrote formula(s) for {added} new row(s) in {', '.join(formula_templates)}"
    return summary

def sync_channel(channel_cfg, workbook_path, sheet_name, records):
    target_columns = channel_cfg["target_columns"]
    formula_templates = channel_cfg.get("formula_templates", {})

    wb = openpyxl.load_workbook(workbook_path)  # formulas preserved (data_only not set)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path.name}")
    ws = wb[sheet_name]

    for i, col_name in enumerate(target_columns, start=1):
        ws.cell(row=1, column=i).value = col_name

    summary = sync_append_dedupe(ws, records, target_columns, formula_templates)

    wb.save(workbook_path)
    wb.close()
    return summary

# =============================================================================
# ERROR EMAIL
# =============================================================================

def send_error_email(problems):
    if not EMAIL_ENABLED or not problems:
        return
    if not EMAIL_FROM or not EMAIL_APP_PASSWORD:
        print(
            "\n(!) Skipping error email: set MONTHLY_TARGET_EMAIL_FROM and "
            "MONTHLY_TARGET_EMAIL_APP_PASSWORD environment variables to enable it."
        )
        return

    body = "The Monthly Target Update run hit the following issue(s):\n\n" + "\n".join(
        f"- {p}" for p in problems
    )

    msg = EmailMessage()
    msg["Subject"] = f"{EMAIL_SUBJECT_PREFIX} {len(problems)} issue(s) on {dt.date.today().isoformat()}"
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.set_content(body)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(EMAIL_FROM, EMAIL_APP_PASSWORD)
            server.send_message(msg)
        print(f"\n-> error email sent to {EMAIL_TO} ({len(problems)} issue(s)).")
    except Exception as exc:
        print(f"\n(!) FAILED to send error email: {exc}")

# =============================================================================
# MAIN
# =============================================================================

def process_period(month, year, all_master_rows, master_headers, is_grace_catchup, problems):
    label = f"{month_name(month, year)} {year}" + (" (grace catch-up)" if is_grace_catchup else "")
    folder = month_folder(month, year)
    print(f"\n--- {label} : {folder} ---")

    matched, unmatched_files = discover_and_match(folder)

    if not matched and not unmatched_files:
        print("  no files found (folder missing or empty) -- nothing to do.")
        return

    for f in unmatched_files:
        msg = f"[{label}] Found file '{f.name}' with no matching channel profile -> skipped."
        print(f"  {msg}")
        problems.append(msg)

    channel_names_with_own_file = {
        name for name, (_f, profile) in matched.items() if not profile.get("is_catchall")
    }

    master_rows = [r for r in all_master_rows if (r["_month"], r["_year"]) == (month, year)]
    sheet_name = sheet_name_for(month, year)

    for profile_name, (file_path, profile) in matched.items():
        try:
            if profile.get("is_catchall"):
                filter_cfg = resolve_catchall_filter(
                    profile, channel_names_with_own_file - {profile_name}
                )
                channel_cfg = {**profile, "filter": filter_cfg}
            else:
                channel_cfg = profile

            records = project_channel_records(master_rows, master_headers, channel_cfg)
            summary = sync_channel(channel_cfg, file_path, sheet_name, records)
            print(f"  [{profile_name}] {len(records)} matching row(s) -> {summary}.")
        except Exception as exc:
            msg = f"[{label}] {profile_name} FAILED: {exc}"
            print(f"  {msg}")
            problems.append(msg)

def main():
    problems = []
    today = SIMULATED_TODAY or dt.date.today()
    periods = active_periods(today)

    print(f"Today: {today.isoformat()}  |  Active period(s): "
          + ", ".join(f"{month_name(m, y)} {y}{' (grace)' if g else ''}" for m, y, g in periods))

    try:
        all_master_rows, master_headers = load_master_rows_for_periods(periods)
    except Exception as exc:
        msg = f"FATAL: could not read Sales Master: {exc}"
        print(msg)
        send_error_email([msg])
        return

    print(f"  -> {len(all_master_rows)} Sales Master row(s) loaded across active period(s).")

    for month, year, is_grace_catchup in periods:
        process_period(month, year, all_master_rows, master_headers, is_grace_catchup, problems)

    print("\nSales Master was opened read-only and was not modified.")

    if problems:
        print(f"\n{len(problems)} issue(s) encountered this run:")
        for p in problems:
            print(f"  - {p}")
    send_error_email(problems)

if __name__ == "__main__":
    main()
